from openpyxl import Workbook
from openpyxl.utils import get_column_letter

wb = Workbook()
ws = wb.active

# Define start and end row/column numbers
start_row, start_col = 2, 2  # B2
end_row, end_col = 3, 4  # D3

# Convert column numbers to letters
start_col_letter = get_column_letter(start_col)
end_col_letter = get_column_letter(end_col)

# Merge cells using row and column numbers
ws.merge_cells(f"{start_col_letter}{start_row}:{end_col_letter}{end_row}")

# Set value in the merged cell
ws[f"{start_col_letter}{start_row}"] = "Merged Cells"

wb.save("merged_cells.xlsx")
