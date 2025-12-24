from openpyxl import Workbook, load_workbook
from openpyxl.styles import PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from tabulate import tabulate

file_pos = [] #[start_row, start_col, end_row, end_col]

# ---------------------------------------------------------------------------------------------------- #
# FUNCTIONS
def collect_data(file_name):
    workbook = load_workbook(filename=file_name)
    sheet = workbook.active

    data_1 = {}
    for data in sheet.iter_rows(min_row = sheet.min_row, max_row = sheet.min_row, values_only = True):
        for each in data:
            data_1[each] = []

    for i in range(1,sheet.max_column + 1):
        result = []
        for row in sheet.iter_cols(min_col = i, max_col = i, min_row = sheet.min_row + 1, values_only = True):
            result.append(row)

        result = list(result[0])
        #print("res : ",result)
        #print(i - 1, list(data_1.keys())[0][i - 1])
        data_1[list(data_1.keys())[i - 1]] = result

    file_pos.append([sheet.min_row, sheet.min_column, sheet.max_row, sheet.min_column])
    #print(data_1)
    return data_1

def compare_table(dict1, dict2):
    differences = []

    all_keys = set(dict1.keys()).union(set(dict2.keys()))

    for key in all_keys:
        list1 = dict1.get(key, [])
        list2 = dict2.get(key, [])

        max_len = max(len(list1), len(list2))
        row_diff = []

        for i in range(max_len):
            val1 = list1[i] if i < len(list1) else None
            val2 = list2[i] if i < len(list2) else None

            if val1 != val2:
                row_diff.append((i, val1, val2))

        if row_diff:
            differences.append([key, row_diff])

    #print("diff :")
    #print(differences)
    #print()
    return differences

def show_table(dic):
    for i in range(len(dic)):
        width = (i + 1) * 4
        print(f'{"|":<1}{list(dic.keys())[i]:^{width}}{"|":>1}',end = '')

    print()
    for i in range(len(dic[list(dic.keys())[0]])):
        for j in range(len(dic)):
            width = (j + 1) * 4
            print(f'{"|":<1}{dic[list(dic.keys())[j]][i]:^{width}}{"|":>1}',end = '')
        print()
    print()


# ---------------------------------------------------------------------------------------------------- #
# INITIATE TABLES
data_sets = [collect_data("file1.xlsx"), collect_data("file2.xlsx")]
diff_table = compare_table(data_sets[0], data_sets[1])

# Format for tabulate display
formatted_diff = []
for key, diffs in diff_table:
    for index, old_val, new_val in diffs:
        formatted_diff.append([key, index, old_val, new_val])

print(formatted_diff)
print(tabulate(formatted_diff, headers=["Key", "Index", "Old Value", "New Value"], tablefmt="grid"))
# ---------------------------------------------------------------------------------------------------- #
# EDIT EXCEL FILE

#[start_row, start_col, end_row, end_col]
def excelification(dic, diff, pos, tipe):
    wb = Workbook()
    ws = wb.active

    arr = list(dic.keys())
    thin_border = Border(left=Side(style="thin"), right=Side(style="thin"),
                     top=Side(style="thin"), bottom=Side(style="thin"))
    
    for i in range(len(arr)):
        ws.cell(row = pos[0], column = pos[1] + i, value = arr[i]).alignment = Alignment(horizontal="center", vertical="center")
        ws.cell(row = pos[0], column = pos[1] + i, value = arr[i]).border = thin_border
        ws.cell(row = pos[0], column = pos[1] + i, value = arr[i]).fill = PatternFill(start_color="ADD8E6", fill_type="solid")
        col_letter = get_column_letter(i + 1)
        ws.column_dimensions[col_letter].width = 20 

    mod = 0
    for key, element in dic.items():
        current_col = pos[1] + mod
        mod2 = 1
        for each in element:
            current_row = pos[0] + mod2
            color = None
            for dif in diff:
                if(key == dif[0]):
                    if(tipe == 1 and dif[2] == each):
                        if(dif[3] is None):
                            color = "90EE90"
                        else:
                            color = "FDDA0D"
                    elif(tipe == 2 and dif[3] == each):
                        if(dif[2] is None):
                            color = "90EE90"
                        else:
                            color = "FDDA0D"

            ws.cell(row = current_row, column = current_col, value = each).alignment = Alignment(horizontal="center", vertical="center")
            ws.cell(row = current_row, column = current_col, value = each).border = thin_border
            if color:
                ws.cell(row = current_row, column = current_col, value = each).fill = PatternFill(start_color=color, fill_type="solid")
            mod2 += 1
        mod += 1

    wb.save("Result" + str(tipe) + ".xlsx")

    
excelification(data_sets[0], formatted_diff, file_pos[0], 1)
excelification(data_sets[1], formatted_diff, file_pos[1], 2)
    

    
