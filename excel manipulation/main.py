from openpyxl import Workbook, load_workbook
from openpyxl.styles import PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from tabulate import tabulate

# ---------------------------------------------------------------------------------------------------- #
# SET UPS
file_pos = [] #[start_row, start_col, end_row, end_col]


# colors
red = 'FF0000'
green = '90EE90'
yellow = 'FDDA0D'

# FUNCTIONS
def collect_data(file_name):
    workbook = load_workbook(filename=file_name)
    sheet = workbook.active

    data_1 = {}
    for data in sheet.iter_rows(min_row = sheet.min_row, max_row = sheet.min_row, values_only = True):
        for each in data:
            data_1[each] = []

    for i in range(1,sheet.max_column + 1):
        result = []
        for row in sheet.iter_cols(min_col = i, max_col = i, min_row = sheet.min_row + 1, values_only = True):
            result.append(row)

        result = list(result[0])
        #print("res : ",result)
        #print(i - 1, list(data_1.keys())[0][i - 1])
        data_1[list(data_1.keys())[i - 1]] = result

    file_pos.append([sheet.min_row, sheet.min_column, sheet.max_row, sheet.min_column])
    #print(data_1)
    return data_1

def compare_table(dict1, dict2):
    differences = []

    all_keys = set(dict1.keys()).union(set(dict2.keys()))

    for key in all_keys:
        list1 = dict1.get(key, [])
        list2 = dict2.get(key, [])

        max_len = max(len(list1), len(list2))
        row_diff = []

        for i in range(max_len):
            val1 = list1[i] if i < len(list1) else None
            val2 = list2[i] if i < len(list2) else None

            if val1 != val2:
                row_diff.append((i, val1, val2))

        if row_diff:
            differences.append([key, row_diff])

    #print("diff :")
    #print(differences)
    #print()
    return differences

def show_table(dic):
    for i in range(len(dic)):
        width = (i + 1) * 4
        print(f'{"|":<1}{list(dic.keys())[i]:^{width}}{"|":>1}',end = '')

    print()
    for i in range(len(dic[list(dic.keys())[0]])):
        for j in range(len(dic)):
            width = (j + 1) * 4
            print(f'{"|":<1}{dic[list(dic.keys())[j]][i]:^{width}}{"|":>1}',end = '')
        print()
    print()


# ---------------------------------------------------------------------------------------------------- #
# INITIATE TABLES
data_sets = [collect_data("file1.xlsx"), collect_data("file2.xlsx")]
diff_table = compare_table(data_sets[0], data_sets[1])

# Format for tabulate display
formatted_diff = []
for key, diffs in diff_table:
    for index, old_val, new_val in diffs:
        formatted_diff.append([key, index, old_val, new_val])

print(formatted_diff)
print(tabulate(formatted_diff, headers=["Key", "Index", "Old Value", "New Value"], tablefmt="grid"))
# ---------------------------------------------------------------------------------------------------- #
# EDIT EXCEL FILE

#[start_row, start_col, end_row, end_col]
def excelification(dic1, dic2, diff, pos):
    wb = Workbook()
    ws = wb.active

    arr = set(dic1.keys()) | set(dic2.keys())
    thin_border = Border(left=Side(style="thin"), right=Side(style="thin"),
                     top=Side(style="thin"), bottom=Side(style="thin"))

    for i in range(len(arr)):
        start_row, start_col = pos[0], pos[1] + i * 2  # B2
        end_row, end_col = pos[0], pos[1] + i * 2 + 1  # B3

        # Convert column numbers to letters
        start_col_letter = get_column_letter(start_col)
        end_col_letter = get_column_letter(end_col)

        ws.merge_cells(f'{start_col_letter}{start_row}:{end_col_letter}{end_row}')
        ws.cell(row = start_row, column = start_col, value = list(arr)[i]).alignment = Alignment(horizontal="center", vertical="center")
        ws.cell(row = start_row, column = start_col, value = list(arr)[i]).border = thin_border
        ws.cell(row = start_row, column = start_col, value = list(arr)[i]).fill = PatternFill(start_color="ADD8E6", fill_type="solid")
        col_letter = get_column_letter(i + 1)
        ws.column_dimensions[col_letter].width = 20 
        col_letter = get_column_letter(i + 2)
        ws.column_dimensions[col_letter].width = 20 


    mod = 0
    last_row = 0
    for category in arr:
        current_row, current_col = pos[0] + 1, pos[1] + mod * 2

        ws.cell(row = current_row, column = current_col, value = 'File 1').alignment = Alignment(horizontal = 'center', vertical = 'center')
        ws.cell(row = current_row, column = current_col + 1, value = 'File 2').alignment = Alignment(horizontal = 'center', vertical = 'center')

        ws.cell(row = current_row, column = current_col).border = thin_border
        ws.cell(row = current_row, column = current_col).fill = PatternFill(start_color="C04000", fill_type="solid")
        ws.cell(row = current_row, column = current_col + 1).border = thin_border
        ws.cell(row = current_row, column = current_col + 1).fill = PatternFill(start_color="C04000", fill_type="solid")

        mod += 1
        last_row = current_row + 4
        for row in range(max(len(dic1[category]), len(dic2[category]))):
            current_row += 1
            color1 = None
            color2 = None
            for dif in diff:
                if(category == dif[0] and row == dif[1]):
                    if(dif[2] is None):
                        color1 = red
                        color2 = green
                    elif(dif[3] is None):
                        color1 = green
                        color2 = red
                    else:
                        color1 = yellow
                        color2 = yellow

            try:
                value1 = dic1[category][row]
            except:
                value1 = None

            try:
                value2 = dic2[category][row]
            except:
                value2 = None


            ws.cell(row = current_row, column = current_col, value = value1).alignment = Alignment(horizontal="center", vertical="center")
            ws.cell(row = current_row, column = current_col + 1, value = value2).alignment = Alignment(horizontal="center", vertical="center")
            ws.cell(row = current_row, column = current_col, value = value1).border = thin_border
            ws.cell(row = current_row, column = current_col + 1, value = value2).border = thin_border
            
            if(color1 and color2):
                ws.cell(row = current_row, column = current_col).fill = PatternFill(start_color=color1, fill_type="solid")
                ws.cell(row = current_row, column = current_col + 1).fill = PatternFill(start_color=color2, fill_type="solid")

    last_row += 5
    pos[1] += 1
    # Information for reader
    ws.cell(row = last_row, column = pos[1], value = "Information").alignment = Alignment(horizontal="center", vertical="center")
    ws.cell(row = last_row + 1, column = pos[1]).fill = PatternFill(start_color = red, fill_type = "solid")
    ws.cell(row = last_row + 2, column = pos[1]).fill = PatternFill(start_color = green, fill_type = "solid")
    ws.cell(row = last_row + 3, column = pos[1]).fill = PatternFill(start_color = yellow, fill_type = "solid")
    
    ws.cell(row = last_row + 1, column = pos[1]).border = thin_border
    ws.cell(row = last_row + 2, column = pos[1]).border = thin_border
    ws.cell(row = last_row + 3, column = pos[1]).border = thin_border

    ws.cell(row = last_row + 1, column = pos[1] + 1,value = "Not found").alignment = Alignment(horizontal="center", vertical="center")
    ws.cell(row = last_row + 2, column = pos[1] + 1,value = "Found").alignment = Alignment(horizontal="center", vertical="center")
    ws.cell(row = last_row + 3, column = pos[1] + 1,value = "Different value").alignment = Alignment(horizontal="center", vertical="center")
    
    wb.save("Result.xlsx")

    
excelification(data_sets[0], data_sets[1], formatted_diff, file_pos[0])
    

    
